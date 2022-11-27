#packages
from flask import Flask, jsonify, request, redirect, render_template, url_for, jsonify, make_response, send_from_directory
from flask_pymongo import PyMongo
from bson.objectid import ObjectId
from datetime import date, datetime, timedelta
# import locale
# locale.setlocale(locale.LC_ALL, 'en_IN.utf8')
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
import os
import calendar
import decimal

#app
app = Flask(__name__)
app.config["MONGO_URI"] = "mongodb+srv://iocofficetool:iocofficetool@cluster0.dnm3eca.mongodb.net/ioc?retryWrites=true&w=majority"
mongo = PyMongo(app)

#collections
projects = mongo.db.projects
villages = mongo.db.villages
records = mongo.db.records

def currency_in_indian_format(n):
    """ Convert a number (int / float) into indian formatting style """
    d = decimal.Decimal(str(n))

    if d.as_tuple().exponent < -2:
        s = str(n)
    else:
        s = '{0:.2f}'.format(n)

    l = len(s)
    i = l - 1

    res, flag, k = '', 0, 0
    while i >= 0:
        if flag == 0:
            res += s[i]
            if s[i] == '.':
                flag = 1
        elif flag == 1:
            k += 1
            res += s[i]
            if k == 3 and i - 1 >= 0:
                res += ','
                flag = 2
                k = 0
        else:
            k += 1
            res += s[i]
            if k == 2 and i - 1 >= 0:
                res += ','
                flag = 2
                k = 0
        i -= 1

    return res[::-1]

#All the routings in our app will be mentioned here.
#Test
@app.route('/test')
def test():
    return currency_in_indian_format(1000000).split(".")[0]

@app.route('/addProject', methods=['POST'])
def addProject():
    session_id = request.cookies.get('session_id')
    if(session_id != "mjeIJPsatvvs"):
        return redirect(url_for('login'))
        
    project = {"projectName": request.form['projectName']}
    projects.insert_one(project)
    return redirect(url_for('config'))

@app.route('/addVillage', methods=['POST'])
def addVillage():
    session_id = request.cookies.get('session_id')
    if(session_id != "mjeIJPsatvvs"):
        return redirect(url_for('login'))
        
    village = {"projectID": ObjectId(request.form['projectID']), "villageName": request.form['villageName']}
    villages.insert_one(village)

    projects_data = projects.find()
    projects_output = [{'projectID': project['_id'], 'projectName' : project['projectName']} for project in projects_data]
    projects_count = len(projects_output)

    villages_data = villages.find({"projectID": ObjectId(request.form['projectID'])})
    villages_output = [{'villageID': village['_id'], 'villageName' : village['villageName']} for village in villages_data]
    villages_count = len(villages_output)

    projectName = ""
    for project in projects_output:
        if(ObjectId(request.form['projectID']) == project['projectID']):
            projectName = project['projectName']
            break
    return render_template('config.html', projects_output=projects_output, projects_count=projects_count, villages_output=villages_output, villages_count=villages_count, display_villages=True, projectID=request.form['projectID'], projectIDObj=ObjectId(request.form['projectID']), projectName=projectName)

@app.route('/addRecord', methods=['POST'])
def addRecord():
    session_id = request.cookies.get('session_id')
    if(session_id != "mjeIJPsatvvs"):
        return redirect(url_for('login'))

    records_count = 0
    if(request.form['transactionType'] == "DISBURSAL"):
        records_data = records.find({"chequeNumber": request.form['chequeNumber']})
        records_output = [{'recordID': str(record['_id'])} for record in records_data]
        records_count = len(records_output)
    isChequeNumberExists = "N"
    addAmountAria = True
    if(records_count > 0):
        isChequeNumberExists = "Y"
        addAmountAria = False
    else:
        record = []
        if(request.form['transactionType'] == "DEPOSIT"):
            record = {"projectID": ObjectId(request.form['projectID']), "villageID": ObjectId(request.form['villageID']), "landOwner": "", "type": request.form['type'], "date": request.form['date'], "amount": int(request.form['amount']), "chequeNumber": "", "serialNumber": "", "transactionType": request.form['transactionType'], "remarks": ""}
        else:
            record = {"projectID": ObjectId(request.form['projectID']), "villageID": ObjectId(request.form['villageID']), "landOwner": request.form['landOwner'], "type": request.form['type'], "date": request.form['date'], "amount": int(request.form['amount']), "chequeNumber": request.form['chequeNumber'], "serialNumber": request.form['serialNumber'], "transactionType": request.form['transactionType'], "remarks": request.form['remarks']}
        records.insert_one(record)
    
    transactionType = request.form['filterTransactionType']
    transactionTypeArray = []
    if(transactionType == "ALL"):
        transactionTypeArray = ["DEPOSIT","DISBURSAL"]
    elif(transactionType == "DEPOSIT"):
        transactionTypeArray = ["DEPOSIT"]
    else:
        transactionTypeArray = ["DISBURSAL"]

    type = request.form['filterType']
    typeArray = []
    if(type == "ALL"):
        typeArray = ["Land","Crop"]
    elif(type == "Land"):
        typeArray = ["Land"]
    else:
        typeArray = ["Crop"]

    fromDate = str(date.min)
    if(request.form['filterFromDate']):
        fromDate = request.form['filterFromDate']
    toDate = str(date.max)
    if(request.form['filterToDate']):
        toDate = request.form['filterToDate']

    records_data = 0
    projectID = request.form['filterProjectID']
    villageID = request.form['filterVillageID']
    matchQuery = {}
    if(projectID == "ALL"):
        matchQuery = {
                        "type": { "$in": typeArray },
                        "transactionType": { "$in": transactionTypeArray },
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ]
                    }
    elif(villageID == "ALL"):
        matchQuery = {
                        "projectID": ObjectId(projectID),
                        "type": { "$in": typeArray },
                        "transactionType": { "$in": transactionTypeArray },
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ]
                    }
    else:
        matchQuery = {
                        "projectID": ObjectId(projectID),
                        "villageID": ObjectId(villageID),
                        "type": { "$in": typeArray },
                        "transactionType": { "$in": transactionTypeArray },
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ]
                    }
    records_data = records.aggregate([
        {
            "$match": matchQuery
        },
        {
            "$lookup":
                {
                    "from": "projects",
                    "localField": "projectID",
                    "foreignField": "_id",
                    "as": "project"
            
                }
        },
        {
            "$unwind":
                {
                    "path": "$project"
                }
        },
        {
            "$lookup":
                {
                    "from": "villages",
                    "localField": "villageID",
                    "foreignField": "_id",
                    "as": "village"
                }
        },
        {
            "$unwind":
                {
                    "path": "$village"
                }
        },
        {
            "$addFields":
                {
                    "remarks": { "$ifNull": ["$remarks", ""] }
                }
        },
        {
            "$sort":
                {
                    "date": 1,
                    "chequeNumber":1
                }
        }
    ])

    records_output = [{'recordID': record['_id'], 'project' : record['project'], 'village' : record['village'], 'landOwner' : record['landOwner'], 'type' : record['type'], 'date' : datetime.strptime(record['date'], '%Y-%m-%d').strftime("%d-%m-%Y"), 'editDate' : record['date'], 'amount' : currency_in_indian_format(record['amount']).split(".")[0], 'intAmount' : record['amount'], 'chequeNumber' : record['chequeNumber'], 'serialNumber' : record['serialNumber'], 'transactionType' : record['transactionType'], 'remarks' : record['remarks']} for record in records_data]
    records_count = len(records_output)
    total_records_count = records.count_documents({})

    projects_data = projects.find()
    projects_output = [{'projectID': str(project['_id']), 'projectName' : project['projectName']} for project in projects_data]
    projects_count = len(projects_output)

    villages_output = []
    villages_count = 0
    if(projects_count != 0):
        villages_data = villages.find({"projectID": ObjectId(request.form['projectID'])})
        villages_output = [{'villageID': str(village['_id']), 'villageName' : village['villageName']} for village in villages_data]
        villages_count = len(villages_output)

    filter_villages_output = []
    filter_villages_count = 0
    if(projectID != "ALL"):
        filter_villages_data = villages.find({"projectID": ObjectId(projectID)})
        filter_villages_output = [{'villageID': str(village['_id']), 'villageName' : village['villageName']} for village in filter_villages_data]
        filter_villages_count = len(filter_villages_output)

    sum = 0
    for record in records_output:
        if(transactionType == "ALL"):
            if(record['transactionType'] == "DEPOSIT"):
                sum += record['intAmount']
            elif(record['transactionType'] == "DISBURSAL"):
                sum -= record['intAmount']
        else:
            sum += record['intAmount']
    sum = currency_in_indian_format(sum).split(".")[0]
    addAmount = currency_in_indian_format(int(request.form['amount'])).split(".")[0]
    sum_text = ""
    if(transactionType == "ALL"):
        sum_text = "DEPOSIT - DISBURSAL"
    elif(transactionType == "DEPOSIT"):
        sum_text = "TOTAL DEPOSIT"
    elif(transactionType == "DISBURSAL"):
        sum_text = "TOTAL DISBURSAL"

    if(fromDate == str(date.min)):
        fromDate = ""
    if(toDate == str(date.max)):
        toDate = ""
    return render_template('statements.html', records_output=records_output, total_records_count=total_records_count, records_count=records_count, display_records=False, projects_output=projects_output, villages_output=villages_output, projects_count=projects_count, villages_count=villages_count, sum=sum, projectID=projectID, villageID=villageID, type=type, transactionType=transactionType, fromDate=fromDate, toDate=toDate, sum_text=sum_text, aria=True, addProjectID = request.form['projectID'], addVillageID = request.form['villageID'], addType = request.form['type'], addDate = request.form['date'], addChequeNumber = request.form['chequeNumber'], addTransactionType = request.form['transactionType'], addAmount = addAmount, filter_villages_output=filter_villages_output, filter_villages_count=filter_villages_count, isChequeNumberExists=isChequeNumberExists, addAmountAria=addAmountAria)

@app.route('/updateProject', methods=['POST'])
def updateProject():
    session_id = request.cookies.get('session_id')
    if(session_id != "mjeIJPsatvvs"):
        return redirect(url_for('login'))
        
    updated_project = {"$set": {'projectName' : request.form['projectName']}}
    filt = {'_id' : ObjectId(request.form['projectID'])}
    projects.update_one(filt, updated_project)
    return redirect(url_for('config'))

@app.route('/updateVillage', methods=['POST'])
def updateVillage():
    session_id = request.cookies.get('session_id')
    if(session_id != "mjeIJPsatvvs"):
        return redirect(url_for('login'))
        
    updated_village = {"$set": {'villageName' : request.form['villageName']}}
    filt = {'_id' : ObjectId(request.form['villageID'])}
    villages.update_one(filt, updated_village)
    
    projects_data = projects.find()
    projects_output = [{'projectID': project['_id'], 'projectName' : project['projectName']} for project in projects_data]
    projects_count = len(projects_output)

    villages_data = villages.find({"projectID": ObjectId(request.form['projectID'])})
    villages_output = [{'villageID': village['_id'], 'villageName' : village['villageName']} for village in villages_data]
    villages_count = len(villages_output)

    projectName = ""
    for project in projects_output:
        if(ObjectId(request.form['projectID']) == project['projectID']):
            projectName = project['projectName']
            break
    return render_template('config.html', projects_output=projects_output, projects_count=projects_count, villages_output=villages_output, villages_count=villages_count, display_villages=True, projectID=request.form['projectID'], projectIDObj=ObjectId(request.form['projectID']), projectName=projectName)

@app.route('/updateRecord', methods=['POST'])
def updateRecord():
    session_id = request.cookies.get('session_id')
    if(session_id != "mjeIJPsatvvs"):
        return redirect(url_for('login'))

    updated_record = []
    if(request.form['editTransactionType'] == "DEPOSIT"):
        updated_record = {"$set": {"projectID": ObjectId(request.form['editProjectID']), "villageID": ObjectId(request.form['editVillageID']), "landOwner": "", "type": request.form['editType'], "date": request.form['editDate'], "amount": int(request.form['editAmount']), "chequeNumber": "", "serialNumber": "", "transactionType": request.form['editTransactionType'], "remarks": ""}}
    else:
        updated_record = {"$set": {"projectID": ObjectId(request.form['editProjectID']), "villageID": ObjectId(request.form['editVillageID']), "landOwner": request.form['editLandOwner'], "type": request.form['editType'], "date": request.form['editDate'], "amount": int(request.form['editAmount']), "chequeNumber": request.form['editChequeNumber'], "serialNumber": request.form['editSerialNumber'], "transactionType": request.form['editTransactionType'], "remarks": request.form['editRemarks']}}
    filt = {'_id' : ObjectId(request.form['recordID'])}
    records.update_one(filt, updated_record)

    transactionType = request.form['filterTransactionType']
    transactionTypeArray = []
    if(transactionType == "ALL"):
        transactionTypeArray = ["DEPOSIT","DISBURSAL"]
    elif(transactionType == "DEPOSIT"):
        transactionTypeArray = ["DEPOSIT"]
    else:
        transactionTypeArray = ["DISBURSAL"]

    type = request.form['filterType']
    typeArray = []
    if(type == "ALL"):
        typeArray = ["Land","Crop"]
    elif(type == "Land"):
        typeArray = ["Land"]
    else:
        typeArray = ["Crop"]

    fromDate = str(date.min)
    if(request.form['filterFromDate']):
        fromDate = request.form['filterFromDate']
    toDate = str(date.max)
    if(request.form['filterToDate']):
        toDate = request.form['filterToDate']

    records_data = 0
    projectID = request.form['filterProjectID']
    villageID = request.form['filterVillageID']
    matchQuery = {}
    if(projectID == "ALL"):
        matchQuery = {
                        "type": { "$in": typeArray },
                        "transactionType": { "$in": transactionTypeArray },
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ]
                    }
    elif(villageID == "ALL"):
        matchQuery = {
                        "projectID": ObjectId(projectID),
                        "type": { "$in": typeArray },
                        "transactionType": { "$in": transactionTypeArray },
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ]
                    }
    else:
        matchQuery = {
                        "projectID": ObjectId(projectID),
                        "villageID": ObjectId(villageID),
                        "type": { "$in": typeArray },
                        "transactionType": { "$in": transactionTypeArray },
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ]
                    }
    records_data = records.aggregate([
        {
            "$match": matchQuery
        },
        {
            "$lookup":
                {
                    "from": "projects",
                    "localField": "projectID",
                    "foreignField": "_id",
                    "as": "project"
            
                }
        },
        {
            "$unwind":
                {
                    "path": "$project"
                }
        },
        {
            "$lookup":
                {
                    "from": "villages",
                    "localField": "villageID",
                    "foreignField": "_id",
                    "as": "village"
                }
        },
        {
            "$unwind":
                {
                    "path": "$village"
                }
        },
        {
            "$addFields":
                {
                    "remarks": { "$ifNull": ["$remarks", ""] }
                }
        },
        {
            "$sort":
                {
                    "date": 1,
                    "chequeNumber":1
                }
        }
    ])

    records_output = [{'recordID': record['_id'], 'project' : record['project'], 'village' : record['village'], 'landOwner' : record['landOwner'], 'type' : record['type'], 'date' : datetime.strptime(record['date'], '%Y-%m-%d').strftime("%d-%m-%Y"), 'editDate' : record['date'], 'amount' : currency_in_indian_format(record['amount']).split(".")[0], 'intAmount' : record['amount'], 'chequeNumber' : record['chequeNumber'], 'serialNumber' : record['serialNumber'], 'transactionType' : record['transactionType'], 'remarks' : record['remarks']} for record in records_data]
    records_count = len(records_output)
    total_records_count = records.count_documents({})

    projects_data = projects.find()
    projects_output = [{'projectID': str(project['_id']), 'projectName' : project['projectName']} for project in projects_data]
    projects_count = len(projects_output)

    villages_data = villages.find({"projectID": ObjectId(projects_output[0]['projectID'])})
    villages_output = [{'villageID': str(village['_id']), 'villageName' : village['villageName']} for village in villages_data]
    villages_count = len(villages_output)

    filter_villages_output = []
    filter_villages_count = 0
    if(projectID != "ALL"):
        filter_villages_data = villages.find({"projectID": ObjectId(projectID)})
        filter_villages_output = [{'villageID': str(village['_id']), 'villageName' : village['villageName']} for village in filter_villages_data]
        filter_villages_count = len(filter_villages_output)

    sum = 0
    for record in records_output:
        if(transactionType == "ALL"):
            if(record['transactionType'] == "DEPOSIT"):
                sum += record['intAmount']
            elif(record['transactionType'] == "DISBURSAL"):
                sum -= record['intAmount']
        else:
            sum += record['intAmount']
    sum = currency_in_indian_format(sum).split(".")[0]
    sum_text = ""
    if(transactionType == "ALL"):
        sum_text = "DEPOSIT - DISBURSAL"
    elif(transactionType == "DEPOSIT"):
        sum_text = "TOTAL DEPOSIT"
    elif(transactionType == "DISBURSAL"):
        sum_text = "TOTAL DISBURSAL"

    if(fromDate == str(date.min)):
        fromDate = ""
    if(toDate == str(date.max)):
        toDate = ""
    return render_template('statements.html', records_output=records_output, total_records_count=total_records_count, records_count=records_count, display_records=False, projects_output=projects_output, villages_output=villages_output, projects_count=projects_count, villages_count=villages_count, sum=sum, projectID=projectID, villageID=villageID, type=type, transactionType=transactionType, fromDate=fromDate, toDate=toDate, filter_villages_output=filter_villages_output, filter_villages_count=filter_villages_count, sum_text=sum_text, isChequeNumberExists="N")

@app.route('/deleteProject', methods=['POST'])
def deleteProject():
    session_id = request.cookies.get('session_id')
    if(session_id != "mjeIJPsatvvs"):
        return redirect(url_for('login'))
        
    filt = {'_id' : ObjectId(request.form['projectID'])}
    projects.delete_one(filt)
    return redirect(url_for('config'))


@app.route('/deleteVillage', methods=['POST'])
def deleteVillage():
    session_id = request.cookies.get('session_id')
    if(session_id != "mjeIJPsatvvs"):
        return redirect(url_for('login'))
        
    filt = {'_id' : ObjectId(request.form['villageID'])}
    villages.delete_one(filt)
    
    projects_data = projects.find()
    projects_output = [{'projectID': project['_id'], 'projectName' : project['projectName']} for project in projects_data]
    projects_count = len(projects_output)

    villages_data = villages.find({"projectID": ObjectId(request.form['projectID'])})
    villages_output = [{'villageID': village['_id'], 'villageName' : village['villageName']} for village in villages_data]
    villages_count = len(villages_output)

    projectName = ""
    for project in projects_output:
        if(ObjectId(request.form['projectID']) == project['projectID']):
            projectName = project['projectName']
            break
    return render_template('config.html', projects_output=projects_output, projects_count=projects_count, villages_output=villages_output, villages_count=villages_count, display_villages=True, projectID=request.form['projectID'], projectIDObj=ObjectId(request.form['projectID']), projectName=projectName)

@app.route('/deleteRecord', methods=['POST'])
def deleteRecord():
    session_id = request.cookies.get('session_id')
    if(session_id != "mjeIJPsatvvs"):
        return redirect(url_for('login'))
        
    filt = {'_id' : ObjectId(request.form['recordID'])}
    records.delete_one(filt)
    
    transactionType = request.form['filterTransactionType']
    transactionTypeArray = []
    if(transactionType == "ALL"):
        transactionTypeArray = ["DEPOSIT","DISBURSAL"]
    elif(transactionType == "DEPOSIT"):
        transactionTypeArray = ["DEPOSIT"]
    else:
        transactionTypeArray = ["DISBURSAL"]

    type = request.form['filterType']
    typeArray = []
    if(type == "ALL"):
        typeArray = ["Land","Crop"]
    elif(type == "Land"):
        typeArray = ["Land"]
    else:
        typeArray = ["Crop"]

    fromDate = str(date.min)
    if(request.form['filterFromDate']):
        fromDate = request.form['filterFromDate']
    toDate = str(date.max)
    if(request.form['filterToDate']):
        toDate = request.form['filterToDate']

    records_data = 0
    projectID = request.form['filterProjectID']
    villageID = request.form['filterVillageID']
    matchQuery = {}
    if(projectID == "ALL"):
        matchQuery = {
                        "type": { "$in": typeArray },
                        "transactionType": { "$in": transactionTypeArray },
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ]
                    }
    elif(villageID == "ALL"):
        matchQuery = {
                        "projectID": ObjectId(projectID),
                        "type": { "$in": typeArray },
                        "transactionType": { "$in": transactionTypeArray },
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ]
                    }
    else:
        matchQuery = {
                        "projectID": ObjectId(projectID),
                        "villageID": ObjectId(villageID),
                        "type": { "$in": typeArray },
                        "transactionType": { "$in": transactionTypeArray },
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ]
                    }
    records_data = records.aggregate([
        {
            "$match": matchQuery
        },
        {
            "$lookup":
                {
                    "from": "projects",
                    "localField": "projectID",
                    "foreignField": "_id",
                    "as": "project"
            
                }
        },
        {
            "$unwind":
                {
                    "path": "$project"
                }
        },
        {
            "$lookup":
                {
                    "from": "villages",
                    "localField": "villageID",
                    "foreignField": "_id",
                    "as": "village"
                }
        },
        {
            "$unwind":
                {
                    "path": "$village"
                }
        },
        {
            "$addFields":
                {
                    "remarks": { "$ifNull": ["$remarks", ""] }
                }
        },
        {
            "$sort":
                {
                    "date": 1,
                    "chequeNumber":1
                }
        }
    ])

    records_output = [{'recordID': record['_id'], 'project' : record['project'], 'village' : record['village'], 'landOwner' : record['landOwner'], 'type' : record['type'], 'date' : datetime.strptime(record['date'], '%Y-%m-%d').strftime("%d-%m-%Y"), 'editDate' : record['date'], 'amount' : currency_in_indian_format(record['amount']).split(".")[0], 'intAmount' : record['amount'], 'chequeNumber' : record['chequeNumber'], 'serialNumber' : record['serialNumber'], 'transactionType' : record['transactionType'], 'remarks' : record['remarks']} for record in records_data]
    records_count = len(records_output)
    total_records_count = records.count_documents({})

    projects_data = projects.find()
    projects_output = [{'projectID': str(project['_id']), 'projectName' : project['projectName']} for project in projects_data]
    projects_count = len(projects_output)

    villages_data = villages.find({"projectID": ObjectId(projects_output[0]['projectID'])})
    villages_output = [{'villageID': str(village['_id']), 'villageName' : village['villageName']} for village in villages_data]
    villages_count = len(villages_output)

    filter_villages_output = []
    filter_villages_count = 0
    if(projectID != "ALL"):
        filter_villages_data = villages.find({"projectID": ObjectId(projectID)})
        filter_villages_output = [{'villageID': str(village['_id']), 'villageName' : village['villageName']} for village in filter_villages_data]
        filter_villages_count = len(filter_villages_output)

    sum = 0
    for record in records_output:
        if(transactionType == "ALL"):
            if(record['transactionType'] == "DEPOSIT"):
                sum += record['intAmount']
            elif(record['transactionType'] == "DISBURSAL"):
                sum -= record['intAmount']
        else:
            sum += record['intAmount']
    sum = currency_in_indian_format(sum).split(".")[0]
    sum_text = ""
    if(transactionType == "ALL"):
        sum_text = "DEPOSIT - DISBURSAL"
    elif(transactionType == "DEPOSIT"):
        sum_text = "TOTAL DEPOSIT"
    elif(transactionType == "DISBURSAL"):
        sum_text = "TOTAL DISBURSAL"

    if(fromDate == str(date.min)):
        fromDate = ""
    if(toDate == str(date.max)):
        toDate = ""
    return render_template('statements.html', records_output=records_output, total_records_count=total_records_count, records_count=records_count, display_records=False, projects_output=projects_output, villages_output=villages_output, projects_count=projects_count, villages_count=villages_count, sum=sum, projectID=projectID, villageID=villageID, type=type, transactionType=transactionType, fromDate=fromDate, toDate=toDate, filter_villages_output=filter_villages_output, filter_villages_count=filter_villages_count, sum_text=sum_text, isChequeNumberExists="N")


@app.route("/")
def statements():
    session_id = request.cookies.get('session_id')
    if(session_id != "mjeIJPsatvvs"):
        return redirect(url_for('login'))

    records_data = records.aggregate([
        {
            "$lookup":
                {
                    "from": "projects",
                    "localField": "projectID",
                    "foreignField": "_id",
                    "as": "project"
            
                }
        },
        {
            "$unwind":
                {
                    "path": "$project"
                }
        },
        {
            "$lookup":
                {
                    "from": "villages",
                    "localField": "villageID",
                    "foreignField": "_id",
                    "as": "village"
                }
        },
        {
            "$unwind":
                {
                    "path": "$village"
                }
        },
        {
            "$addFields":
                {
                    "remarks": { "$ifNull": ["$remarks", ""] }
                }
        },
        {
            "$sort":
                {
                    "date": 1,
                    "chequeNumber":1
                }
        }
    ])
    
    records_output = [{'recordID': record['_id'], 'project' : record['project'], 'village' : record['village'], 'landOwner' : record['landOwner'], 'type' : record['type'], 'date' : datetime.strptime(record['date'], '%Y-%m-%d').strftime("%d-%m-%Y"), 'editDate' : record['date'], 'amount' : currency_in_indian_format(record['amount']).split(".")[0], 'intAmount' : record['amount'], 'chequeNumber' : record['chequeNumber'], 'serialNumber' : record['serialNumber'], 'transactionType' : record['transactionType'], 'remarks' : record['remarks']} for record in records_data]
    total_records_count = len(records_output)

    projects_data = projects.find()
    projects_output = [{'projectID': str(project['_id']), 'projectName' : project['projectName']} for project in projects_data]
    projects_count = len(projects_output)

    villages_output = []
    villages_count = 0
    if(projects_count != 0):
        villages_data = villages.find({"projectID": ObjectId(projects_output[0]['projectID'])})
        villages_output = [{'villageID': str(village['_id']), 'villageName' : village['villageName']} for village in villages_data]
        villages_count = len(villages_output)

    sum = 0
    for record in records_output:
        if(record['transactionType'] == "DEPOSIT"):
            sum += record['intAmount']
        elif(record['transactionType'] == "DISBURSAL"):
            sum -= record['intAmount']
    sum = currency_in_indian_format(sum).split(".")[0]
    sum_text = "DEPOSIT - DISBURSAL"
    return render_template('statements.html', records_output=records_output, total_records_count=total_records_count, records_count=total_records_count, display_records=False, projects_output=projects_output, villages_output=villages_output, projects_count=projects_count, villages_count=villages_count, sum=sum, projectID="ALL", villageID="ALL", type="ALL", transactionType="ALL", fromDate="", toDate="", sum_text=sum_text, aria=False, addProjectID = "", addVillageID = "", addLandOwner = "", addType = "Land", addDate = "", addChequeNumber = "", addSerialNumber = "", addTransactionType = "DEPOSIT", addAmount = "", isChequeNumberExists="N")

@app.route("/viewStatement", methods=['POST'])
def viewStatement():
    session_id = request.cookies.get('session_id')
    if(session_id != "mjeIJPsatvvs"):
        return redirect(url_for('login'))
        
    transactionType = request.form['filterTransactionType']
    transactionTypeArray = []
    if(transactionType == "ALL"):
        transactionTypeArray = ["DEPOSIT","DISBURSAL"]
    elif(transactionType == "DEPOSIT"):
        transactionTypeArray = ["DEPOSIT"]
    else:
        transactionTypeArray = ["DISBURSAL"]

    type = request.form['filterType']
    typeArray = []
    if(type == "ALL"):
        typeArray = ["Land","Crop"]
    elif(type == "Land"):
        typeArray = ["Land"]
    else:
        typeArray = ["Crop"]

    fromDate = str(date.min)
    if(request.form['filterFromDate']):
        fromDate = request.form['filterFromDate']
    toDate = str(date.max)
    if(request.form['filterToDate']):
        toDate = request.form['filterToDate']

    records_data = 0
    projectID = request.form['filterProjectID']
    villageID = request.form['filterVillageID']
    matchQuery = {}
    if(projectID == "ALL"):
        matchQuery = {
                        "type": { "$in": typeArray },
                        "transactionType": { "$in": transactionTypeArray },
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ]
                    }
    elif(villageID == "ALL"):
        matchQuery = {
                        "projectID": ObjectId(projectID),
                        "type": { "$in": typeArray },
                        "transactionType": { "$in": transactionTypeArray },
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ]
                    }
    else:
        matchQuery = {
                        "projectID": ObjectId(projectID),
                        "villageID": ObjectId(villageID),
                        "type": { "$in": typeArray },
                        "transactionType": { "$in": transactionTypeArray },
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ]
                    }
    records_data = records.aggregate([
        {
            "$match": matchQuery
        },
        {
            "$lookup":
                {
                    "from": "projects",
                    "localField": "projectID",
                    "foreignField": "_id",
                    "as": "project"
            
                }
        },
        {
            "$unwind":
                {
                    "path": "$project"
                }
        },
        {
            "$lookup":
                {
                    "from": "villages",
                    "localField": "villageID",
                    "foreignField": "_id",
                    "as": "village"
                }
        },
        {
            "$unwind":
                {
                    "path": "$village"
                }
        },
        {
            "$addFields":
                {
                    "remarks": { "$ifNull": ["$remarks", ""] }
                }
        },
        {
            "$sort":
                {
                    "date": 1,
                    "chequeNumber": 1
                }
        }
    ])

    records_output = [{'recordID': record['_id'], 'project' : record['project'], 'village' : record['village'], 'landOwner' : record['landOwner'], 'type' : record['type'], 'date' : datetime.strptime(record['date'], '%Y-%m-%d').strftime("%d-%m-%Y"), 'editDate' : record['date'], 'amount' : currency_in_indian_format(record['amount']).split(".")[0], 'intAmount' : record['amount'], 'chequeNumber' : record['chequeNumber'], 'serialNumber' : record['serialNumber'], 'transactionType' : record['transactionType'], 'remarks' : record['remarks']} for record in records_data]
    records_count = len(records_output)
    total_records_count = records.count_documents({})

    projects_data = projects.find()
    projects_output = [{'projectID': str(project['_id']), 'projectName' : project['projectName']} for project in projects_data]
    projects_count = len(projects_output)

    villages_data = villages.find({"projectID": ObjectId(projects_output[0]['projectID'])})
    villages_output = [{'villageID': str(village['_id']), 'villageName' : village['villageName']} for village in villages_data]
    villages_count = len(villages_output)

    filter_villages_output = []
    filter_villages_count = 0
    if(projectID != "ALL"):
        filter_villages_data = villages.find({"projectID": ObjectId(projectID)})
        filter_villages_output = [{'villageID': str(village['_id']), 'villageName' : village['villageName']} for village in filter_villages_data]
        filter_villages_count = len(filter_villages_output)

    sum = 0
    for record in records_output:
        if(transactionType == "ALL"):
            if(record['transactionType'] == "DEPOSIT"):
                sum += record['intAmount']
            elif(record['transactionType'] == "DISBURSAL"):
                sum -= record['intAmount']
        else:
            sum += record['intAmount']
    sum = currency_in_indian_format(sum).split(".")[0]
    sum_text = ""
    if(transactionType == "ALL"):
        sum_text = "DEPOSIT - DISBURSAL"
    elif(transactionType == "DEPOSIT"):
        sum_text = "TOTAL DEPOSIT"
    elif(transactionType == "DISBURSAL"):
        sum_text = "TOTAL DISBURSAL"

    if(fromDate == str(date.min)):
        fromDate = ""
    if(toDate == str(date.max)):
        toDate = ""
    return render_template('statements.html', records_output=records_output, total_records_count=total_records_count, records_count=records_count, display_records=False, projects_output=projects_output, villages_output=villages_output, projects_count=projects_count, villages_count=villages_count, sum=sum, projectID=projectID, villageID=villageID, type=type, transactionType=transactionType, fromDate=fromDate, toDate=toDate, filter_villages_output=filter_villages_output, filter_villages_count=filter_villages_count, sum_text=sum_text, isChequeNumberExists="N")

@app.route("/generateExcel", methods=['POST'])
def generateExcel():
    session_id = request.cookies.get('session_id')
    if(session_id != "mjeIJPsatvvs"):
        return redirect(url_for('login'))

    fromDate = request.form['fromDate']
    toDate = request.form['toDate']
    generateType = request.form['generateType']
    noDataMonths = request.form['noDataMonths']
    previousFromDate = datetime(2018, 4, 1).strftime("%Y-%m-%d")
    if(request.form['previousFromDate']):
        previousFromDate = request.form['previousFromDate']
    previousToDate = (datetime.strptime(fromDate, '%Y-%m-%d') - timedelta(days=1)).strftime("%Y-%m-%d")
    if(request.form['previousToDate']):
        previousToDate = request.form['previousToDate']
    projectID = request.form['projectID']
    villageID = request.form['villageID']

    matchObj = {}
    projectObj = {}
    groupObj = {}
    totalColumns = 0
    if(generateType == "monthlyDepositStatement"):
        if(projectID == "ALL"):
            matchObj = {
                        "transactionType": "DEPOSIT",
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ]
                    }
        elif(villageID == "ALL"):
            matchObj = {
                        "transactionType": "DEPOSIT",
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ],
                        "projectID": ObjectId(projectID)
                    }
        else:
            matchObj = {
                        "transactionType": "DEPOSIT",
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ],
                        "projectID": ObjectId(projectID),
                        "villageID": ObjectId(villageID)
                    }
    elif(generateType == "monthlyDepositAbstract" or generateType == "monthlyDepositStatementVillageWise"):
        matchObj = {
                    "transactionType": "DEPOSIT",
                    "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ]
                }
    elif(generateType == "monthlyDisbursalStatement"):
        if(projectID == "ALL"):
            matchObj = {
                        "transactionType": "DISBURSAL",
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ]
                    }
        elif(villageID == "ALL"):
            matchObj = {
                        "transactionType": "DISBURSAL",
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ],
                        "projectID": ObjectId(projectID)
                    }
        else:
            matchObj = {
                        "transactionType": "DISBURSAL",
                        "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ],
                        "projectID": ObjectId(projectID),
                        "villageID": ObjectId(villageID)
                    }
    elif(generateType == "monthlyDisbursalAbstract" or generateType == "monthlyDisbursalStatementVillageWise"):
        matchObj = {
                    "transactionType": "DISBURSAL",
                    "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ]
                }
    elif(generateType == "totalDepositStatement" or generateType == "totalDepositAbstract"):
        matchObj = {
                    "transactionType": "DEPOSIT",
                    "$and": [ { "date": { "$gte": previousFromDate, "$lte": toDate } } ]
                }
    elif(generateType == "totalDisbursalStatement" or generateType == "totalDisbursalAbstract"):
        matchObj = {
                    "transactionType": "DISBURSAL",
                    "$and": [ { "date": { "$gte": previousFromDate, "$lte": toDate } } ]
                }
    elif(generateType == "balanceSheetStatement" or generateType == "balanceSheetAbstract"):
        matchObj = {
                    "$and": [ { "date": { "$gte": fromDate, "$lte": toDate } } ]
                }
    if(generateType == "monthlyDepositStatement"):
        projectObj = {
                    "date": 1,
                    "project": "$project.projectName",
                    "village": "$village.villageName",
                    "land": { "$cond": [ { "$eq": [ "$type", "Land" ] }, "$amount", 0 ] },
                    "crop": { "$cond": [ { "$eq": [ "$type", "Crop" ] }, "$amount", 0 ] },
                    "amount": 1
                }
    elif(generateType == "monthlyDisbursalStatement"):
        projectObj = {
                    "date": 1,
                    "project": "$project.projectName",
                    "village": "$village.villageName",
                    "land": { "$cond": [ { "$eq": [ "$type", "Land" ] }, "$amount", 0 ] },
                    "crop": { "$cond": [ { "$eq": [ "$type", "Crop" ] }, "$amount", 0 ] },
                    "amount": 1,
                    "landOwner": 1,
                    "chequeNumber": 1,
                    "serialNumber": 1,
                    "remarks": { "$ifNull": ["$remarks", ""] }
                }
    elif(generateType == "monthlyDepositAbstract" or generateType == "monthlyDisbursalAbstract"):
        projectObj = {
                    "project": "$project.projectName",
                    "land": { "$cond": [ { "$eq": [ "$type", "Land" ] }, "$amount", 0 ] },
                    "crop": { "$cond": [ { "$eq": [ "$type", "Crop" ] }, "$amount", 0 ] },
                    "amount": 1
                }
    elif(generateType == "monthlyDepositStatementVillageWise" or generateType == "monthlyDisbursalStatementVillageWise"):
        projectObj = {
                    "project": "$project.projectName",
                    "village": "$village.villageName",
                    "land": { "$cond": [ { "$eq": [ "$type", "Land" ] }, "$amount", 0 ] },
                    "crop": { "$cond": [ { "$eq": [ "$type", "Crop" ] }, "$amount", 0 ] },
                    "amount": 1
                }
    elif(generateType == "totalDepositStatement" or generateType == "totalDisbursalStatement"):
        projectObj = {
                    "project": "$project.projectName",
                    "village": "$village.villageName",
                    "pastLand": { "$cond": [ { "$and": [{ "$eq": [ "$type", "Land" ]}, {"$gte": [ "$date", previousFromDate ]}, {"$lte": [ "$date", previousToDate ]}] }, "$amount", 0 ] },
                    "pastCrop": { "$cond": [ { "$and": [{ "$eq": [ "$type", "Crop" ]}, {"$gte": [ "$date", previousFromDate ]}, {"$lte": [ "$date", previousToDate ]}] }, "$amount", 0 ] },
                    "pastTotal": { "$cond": [ { "$and": [{ "$gte": [ "$date", previousFromDate ]}, {"$lte": [ "$date", previousToDate ]}] }, "$amount", 0 ] },
                    "presentLand": { "$cond": [ { "$and": [{ "$eq": [ "$type", "Land" ]}, {"$gte": [ "$date", fromDate ]}, {"$lte": [ "$date", toDate ]}] }, "$amount", 0 ] },
                    "presentCrop": { "$cond": [ { "$and": [{ "$eq": [ "$type", "Crop" ]}, {"$gte": [ "$date", fromDate ]}, {"$lte": [ "$date", toDate ]}] }, "$amount", 0 ] },
                    "presentTotal": { "$cond": [ { "$and": [{"$gte": [ "$date", fromDate ]}, {"$lte": [ "$date", toDate ]}] }, "$amount", 0 ] },
                    "soFarLand": { "$cond": [ { "$eq": [ "$type", "Land" ] }, "$amount", 0 ] },
                    "soFarCrop": { "$cond": [ { "$eq": [ "$type", "Crop" ] }, "$amount", 0 ] },
                    "soFarTotal": "$amount"
                }
    elif(generateType == "totalDepositAbstract" or generateType == "totalDisbursalAbstract"):
        projectObj = {
                    "project": "$project.projectName",
                    "pastLand": { "$cond": [ { "$and": [{ "$eq": [ "$type", "Land" ]}, {"$gte": [ "$date", previousFromDate ]}, {"$lte": [ "$date", previousToDate ]}] }, "$amount", 0 ] },
                    "pastCrop": { "$cond": [ { "$and": [{ "$eq": [ "$type", "Crop" ]}, {"$gte": [ "$date", previousFromDate ]}, {"$lte": [ "$date", previousToDate ]}] }, "$amount", 0 ] },
                    "pastTotal": { "$cond": [ { "$and": [{ "$gte": [ "$date", previousFromDate ]}, {"$lte": [ "$date", previousToDate ]}] }, "$amount", 0 ] },
                    "presentLand": { "$cond": [ { "$and": [{ "$eq": [ "$type", "Land" ]}, {"$gte": [ "$date", fromDate ]}, {"$lte": [ "$date", toDate ]}] }, "$amount", 0 ] },
                    "presentCrop": { "$cond": [ { "$and": [{ "$eq": [ "$type", "Crop" ]}, {"$gte": [ "$date", fromDate ]}, {"$lte": [ "$date", toDate ]}] }, "$amount", 0 ] },
                    "presentTotal": { "$cond": [ { "$and": [{"$gte": [ "$date", fromDate ]}, {"$lte": [ "$date", toDate ]}] }, "$amount", 0 ] },
                    "soFarLand": { "$cond": [ { "$eq": [ "$type", "Land" ] }, "$amount", 0 ] },
                    "soFarCrop": { "$cond": [ { "$eq": [ "$type", "Crop" ] }, "$amount", 0 ] },
                    "soFarTotal": "$amount"
                }
    elif(generateType == "balanceSheetStatement"):
        projectObj = {
                    "project": "$project.projectName",
                    "village": "$village.villageName",
                    "depositLand": { "$cond": [ { "$and": [{ "$eq": [ "$type", "Land" ]}, {"$eq": [ "$transactionType", "DEPOSIT" ]}] }, "$amount", 0 ] },
                    "depositCrop": { "$cond": [ { "$and": [{ "$eq": [ "$type", "Crop" ]}, {"$eq": [ "$transactionType", "DEPOSIT" ]}] }, "$amount", 0 ] },
                    "depositTotal": { "$cond": [ { "$eq": [ "$transactionType", "DEPOSIT" ] }, "$amount", 0 ] },
                    "disbursalLand": { "$cond": [ { "$and": [{ "$eq": [ "$type", "Land" ]}, {"$eq": [ "$transactionType", "DISBURSAL" ]}] }, "$amount", 0 ] },
                    "disbursalCrop": { "$cond": [ { "$and": [{ "$eq": [ "$type", "Crop" ]}, {"$eq": [ "$transactionType", "DISBURSAL" ]}] }, "$amount", 0 ] },
                    "disbursalTotal": { "$cond": [ {"$eq": [ "$transactionType", "DISBURSAL" ] }, "$amount", 0 ] },
                }
    elif(generateType == "balanceSheetAbstract"):
        projectObj = {
                    "project": "$project.projectName",
                    "depositLand": { "$cond": [ { "$and": [{ "$eq": [ "$type", "Land" ]}, {"$eq": [ "$transactionType", "DEPOSIT" ]}] }, "$amount", 0 ] },
                    "depositCrop": { "$cond": [ { "$and": [{ "$eq": [ "$type", "Crop" ]}, {"$eq": [ "$transactionType", "DEPOSIT" ]}] }, "$amount", 0 ] },
                    "depositTotal": { "$cond": [ { "$eq": [ "$transactionType", "DEPOSIT" ] }, "$amount", 0 ] },
                    "disbursalLand": { "$cond": [ { "$and": [{ "$eq": [ "$type", "Land" ]}, {"$eq": [ "$transactionType", "DISBURSAL" ]}] }, "$amount", 0 ] },
                    "disbursalCrop": { "$cond": [ { "$and": [{ "$eq": [ "$type", "Crop" ]}, {"$eq": [ "$transactionType", "DISBURSAL" ]}] }, "$amount", 0 ] },
                    "disbursalTotal": { "$cond": [ {"$eq": [ "$transactionType", "DISBURSAL" ] }, "$amount", 0 ] },
                }
    if(generateType == "monthlyDepositStatement"):
        groupObj = {
                    "_id": { "date": "$date", "project": "$project", "village": "$village"},
                    "date": { "$first": "$date" },
                    "project": { "$first": "$project" },
                    "village": { "$first": "$village" },
                    "land": {"$sum": "$land" },
                    "crop": {"$sum": "$crop" },
                    "total": {"$sum": "$amount" }
                }
    elif(generateType == "monthlyDisbursalStatement"):
        groupObj = {
                    "_id": { "date": "$date", "project": "$project", "village": "$village", "landOwner": "$landOwner", "chequeNumber": "$chequeNumber", "serialNumber": "$serialNumber"},
                    "date": { "$first": "$date" },
                    "project": { "$first": "$project" },
                    "village": { "$first": "$village" },
                    "land": {"$sum": "$land" },
                    "crop": {"$sum": "$crop" },
                    "total": {"$sum": "$amount" },
                    "landOwner": { "$first": "$landOwner" },
                    "chequeNumber": { "$first": "$chequeNumber" },
                    "serialNumber": { "$first": "$serialNumber" },
                    "remarks": { "$first": "$remarks" }
                }
    elif(generateType == "monthlyDepositAbstract" or generateType == "monthlyDisbursalAbstract"):
        groupObj = {
                    "_id": "$project",
                    "project": { "$first": "$project" },
                    "land": {"$sum": "$land" },
                    "crop": {"$sum": "$crop" },
                    "total": {"$sum": "$amount" }
                }
    elif(generateType == "monthlyDepositStatementVillageWise" or generateType == "monthlyDisbursalStatementVillageWise"):
        groupObj = {
                    "_id": { "project": "$project", "village": "$village" },
                    "project": { "$first": "$project" },
                    "village": { "$first": "$village" },
                    "land": {"$sum": "$land" },
                    "crop": {"$sum": "$crop" },
                    "total": {"$sum": "$amount" }
                }
    elif(generateType == "totalDepositStatement" or generateType == "totalDisbursalStatement"):
        groupObj = {
                    "_id": { "project": "$project", "village": "$village" },
                    "project": { "$first": "$project" },
                    "village": { "$first": "$village" },
                    "pastLand": {"$sum": "$pastLand" },
                    "pastCrop": {"$sum": "$pastCrop" },
                    "pastTotal": {"$sum": "$pastTotal" },
                    "presentLand": {"$sum": "$presentLand" },
                    "presentCrop": {"$sum": "$presentCrop" },
                    "presentTotal": {"$sum": "$presentTotal" },
                    "soFarLand": {"$sum": "$soFarLand" },
                    "soFarCrop": {"$sum": "$soFarCrop" },
                    "soFarTotal": {"$sum": "$soFarTotal" }
                }
    elif(generateType == "totalDepositAbstract" or generateType == "totalDisbursalAbstract"):
        groupObj = {
                    "_id": "$project",
                    "project": { "$first": "$project" },
                    "pastLand": {"$sum": "$pastLand" },
                    "pastCrop": {"$sum": "$pastCrop" },
                    "pastTotal": {"$sum": "$pastTotal" },
                    "presentLand": {"$sum": "$presentLand" },
                    "presentCrop": {"$sum": "$presentCrop" },
                    "presentTotal": {"$sum": "$presentTotal" },
                    "soFarLand": {"$sum": "$soFarLand" },
                    "soFarCrop": {"$sum": "$soFarCrop" },
                    "soFarTotal": {"$sum": "$soFarTotal" }
                }
    elif(generateType == "balanceSheetStatement"):
        groupObj = {
                    "_id": { "project": "$project", "village": "$village" },
                    "project": { "$first": "$project" },
                    "village": { "$first": "$village" },
                    "depositLand": {"$sum": "$depositLand" },
                    "depositCrop": {"$sum": "$depositCrop" },
                    "depositTotal": {"$sum": "$depositTotal" },
                    "disbursalLand": {"$sum": "$disbursalLand" },
                    "disbursalCrop": {"$sum": "$disbursalCrop" },
                    "disbursalTotal": {"$sum": "$disbursalTotal" }
                }
    elif(generateType == "balanceSheetAbstract"):
        groupObj = {
                    "_id": "$project",
                    "project": { "$first": "$project" },
                    "depositLand": {"$sum": "$depositLand" },
                    "depositCrop": {"$sum": "$depositCrop" },
                    "depositTotal": {"$sum": "$depositTotal" },
                    "disbursalLand": {"$sum": "$disbursalLand" },
                    "disbursalCrop": {"$sum": "$disbursalCrop" },
                    "disbursalTotal": {"$sum": "$disbursalTotal" }
                }

    if(generateType == "monthlyDepositStatement"):
        if(projectID == "ALL"):
            totalColumns = 7
        elif(villageID == "ALL"):
            totalColumns = 6
        else:
            totalColumns = 5
    elif(generateType == "monthlyDepositAbstract" or generateType == "monthlyDisbursalAbstract" or generateType == "monthlyDepositStatementVillageWise" or generateType == "monthlyDisbursalStatementVillageWise"):
        totalColumns = 5
    elif(generateType == "totalDepositStatement" or generateType == "totalDisbursalStatement" or generateType == "totalDepositAbstract" or generateType == "totalDisbursalAbstract" or generateType == "balanceSheetStatement" or generateType == "balanceSheetAbstract"):
        totalColumns = 11
    elif(generateType == "monthlyDisbursalStatement"):
        if(projectID == "ALL"):
            totalColumns = 11
        elif(villageID == "ALL"):
            totalColumns = 10
        else:
            totalColumns = 9
    records_data = records.aggregate([
        {
            "$match": matchObj
        },
        {
            "$lookup":
                {
                    "from": "projects",
                    "localField": "projectID",
                    "foreignField": "_id",
                    "as": "project"
            
                }
        },
        {
            "$unwind":
                {
                    "path": "$project"
                }
        },
        {
            "$lookup":
                {
                    "from": "villages",
                    "localField": "villageID",
                    "foreignField": "_id",
                    "as": "village"
                }
        },
        {
            "$unwind":
                {
                    "path": "$village"
                }
        },
        {
            "$project": projectObj
        },
        {
            "$group": groupObj
        },
        {
            "$sort":
                {
                    "date": 1,
                    "chequeNumber": 1
                }
        }
    ])

    records_output = []
    if(generateType == "monthlyDepositStatement"):
        records_output = [{'project' : record['project'], 'village' : record['village'], 'date' : datetime.strptime(record['date'], '%Y-%m-%d').strftime("%d-%m-%Y"), 'intLand' : record['land'], 'intCrop' : record['crop'], 'intAmount' : record['total']} for record in records_data]
    elif(generateType == "monthlyDisbursalStatement"):
        records_output = [{'project' : record['project'], 'village' : record['village'], 'date' : datetime.strptime(record['date'], '%Y-%m-%d').strftime("%d-%m-%Y"), 'intLand' : record['land'], 'intCrop' : record['crop'], 'intAmount' : record['total'], 'landOwner' : record['landOwner'], 'chequeNumber' : record['chequeNumber'], 'serialNumber' : record['serialNumber'], 'remarks' : record['remarks']} for record in records_data]
    elif(generateType == "monthlyDepositAbstract" or generateType == "monthlyDisbursalAbstract"):
        records_output = [{'project' : record['project'], 'intLand' : record['land'], 'intCrop' : record['crop'], 'intAmount' : record['total']} for record in records_data]
    elif(generateType == "monthlyDepositStatementVillageWise" or generateType == "monthlyDisbursalStatementVillageWise"):
        records_output = [{
            'project' : record['project'],
            'village' : record['village'],
            'intLand' : record['land'],
            'intCrop' : record['crop'],
            'intAmount' : record['total']
            } for record in records_data]
    elif(generateType == "totalDepositStatement" or generateType == "totalDisbursalStatement"):
        records_output = [{
            'project' : record['project'],
            'village' : record['village'],
            'intPastLand' : record['pastLand'],
            'intPastCrop' : record['pastCrop'],
            'intPastTotal' : record['pastTotal'],
            'intPresentLand' : record['presentLand'],
            'intPresentCrop' : record['presentCrop'],
            'intPresentTotal' : record['presentTotal'],
            'intSoFarLand' : record['soFarLand'],
            'intSoFarCrop' : record['soFarCrop'],
            'intSoFarTotal' : record['soFarTotal']
            } for record in records_data]
    elif(generateType == "totalDepositAbstract" or generateType == "totalDisbursalAbstract"):
        records_output = [{
            'project' : record['project'],
            'intPastLand' : record['pastLand'],
            'intPastCrop' : record['pastCrop'],
            'intPastTotal' : record['pastTotal'],
            'intPresentLand' : record['presentLand'],
            'intPresentCrop' : record['presentCrop'],
            'intPresentTotal' : record['presentTotal'],
            'intSoFarLand' : record['soFarLand'],
            'intSoFarCrop' : record['soFarCrop'],
            'intSoFarTotal' : record['soFarTotal']
            } for record in records_data]
    elif(generateType == "balanceSheetStatement"):
        records_output = [{
            'project' : record['project'],
            'village' : record['village'],
            'intDepositLand' : record['depositLand'],
            'intDepositCrop' : record['depositCrop'],
            'intDepositTotal' : record['depositTotal'],
            'intDisbursalLand' : record['disbursalLand'],
            'intDisbursalCrop' : record['disbursalCrop'],
            'intDisbursalTotal' : record['disbursalTotal'],
            'intBalanceLand' : record['depositLand'] - record['disbursalLand'],
            'intBalanceCrop' : record['depositCrop'] - record['disbursalCrop'],
            'intBalanceTotal' : record['depositTotal'] - record['disbursalTotal']
            } for record in records_data]
    elif(generateType == "balanceSheetAbstract"):
        records_output = [{
            'project' : record['project'],
            'intDepositLand' : record['depositLand'],
            'intDepositCrop' : record['depositCrop'],
            'intDepositTotal' : record['depositTotal'],
            'intDisbursalLand' : record['disbursalLand'],
            'intDisbursalCrop' : record['disbursalCrop'],
            'intDisbursalTotal' : record['disbursalTotal'],
            'intBalanceLand' : record['depositLand'] - record['disbursalLand'],
            'intBalanceCrop' : record['depositCrop'] - record['disbursalCrop'],
            'intBalanceTotal' : record['depositTotal'] - record['disbursalTotal']
            } for record in records_data]
    records_count = len(records_output)

    landSum = 0
    cropSum = 0
    totalSum = 0
    pastLandSum = 0
    pastCropSum = 0
    pastTotalSum = 0
    presentLandSum = 0
    presentCropSum = 0
    presentTotalSum = 0
    soFarLandSum = 0
    soFarCropSum = 0
    soFarTotalSum = 0
    depositLandSum = 0
    depositCropSum = 0
    depositTotalSum = 0
    disbursalLandSum = 0
    disbursalCropSum = 0
    disbursalTotalSum = 0
    balanceLandSum = 0
    balanceCropSum = 0
    balanceTotalSum = 0
    if(generateType == "monthlyDepositStatement" or generateType == "monthlyDisbursalStatement" or generateType == "monthlyDepositAbstract" or generateType == "monthlyDisbursalAbstract" or generateType == "monthlyDepositStatementVillageWise" or generateType == "monthlyDisbursalStatementVillageWise"):
        for record in records_output:
            landSum += record['intLand']
            cropSum += record['intCrop']
            totalSum += record['intAmount']
    elif(generateType == "totalDepositStatement" or generateType == "totalDisbursalStatement" or generateType == "totalDepositAbstract" or generateType == "totalDisbursalAbstract"):
        for record in records_output:
            pastLandSum += record['intPastLand']
            pastCropSum += record['intPastCrop']
            pastTotalSum += record['intPastTotal']
            presentLandSum += record['intPresentLand']
            presentCropSum += record['intPresentCrop']
            presentTotalSum += record['intPresentTotal']
            soFarLandSum += record['intSoFarLand']
            soFarCropSum += record['intSoFarCrop']
            soFarTotalSum += record['intSoFarTotal']
    elif(generateType == "balanceSheetStatement" or generateType == "balanceSheetAbstract"):
        for record in records_output:
            depositLandSum += record['intDepositLand']
            depositCropSum += record['intDepositCrop']
            depositTotalSum += record['intDepositTotal']
            disbursalLandSum += record['intDisbursalLand']
            disbursalCropSum += record['intDisbursalCrop']
            disbursalTotalSum += record['intDisbursalTotal']
            balanceLandSum += record['intBalanceLand']
            balanceCropSum += record['intBalanceCrop']
            balanceTotalSum += record['intBalanceTotal']

    previousFromDate = datetime.strptime(previousFromDate, '%Y-%m-%d').strftime("%d-%m-%Y")
    previousToDate = datetime.strptime(previousToDate, '%Y-%m-%d').strftime("%d-%m-%Y")
    fromDate = datetime.strptime(fromDate, '%Y-%m-%d').strftime("%d-%m-%Y")
    toDate = datetime.strptime(toDate, '%Y-%m-%d').strftime("%d-%m-%Y")

    wb = openpyxl.Workbook() 
    sheet = wb.active

    sheet.append(["STATEMENT OF SRPL PROJECTS - IOCL, THIRUVALLUR"])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=totalColumns)
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    rowNumber = 4
    sumRows = []
    currencyFormat = '#,##0'
    if(generateType == "monthlyDepositStatement"):
        sheet.append(["STATEMENT OF DEPOSIT (DATE WISE) - FROM " + fromDate + " TO " + toDate])
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=totalColumns)
        sheet.cell(row=2, column=1).font = Font(bold=True)
        sheet.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
        if(projectID == "ALL"):
            sheet.append(["Sl. No.", "Date", "Project",  "Village", "Land\nRs.", "Crop\nRs.", "Total\nRs."])
            sheet.append([1,2,3,4,5,6,7])
            for i in range(1,totalColumns+1):
                sheet.cell(row=3, column=i).alignment = Alignment(horizontal="center", vertical="center")
                sheet.cell(row=4, column=i).alignment = Alignment(horizontal="center", vertical="center")
            sheet.print_title_rows = '3:3'
        elif(villageID == "ALL"):
            project = projects.find_one({"_id": ObjectId(projectID)})
            projectName = project['projectName']
            sheet.append(["Project - " + projectName])
            rowNumber += 1
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=totalColumns)
            sheet.cell(row=3, column=1).font = Font(bold=True)
            sheet.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
            sheet.append(["Sl. No.", "Date", "Village", "Land\nRs.", "Crop\nRs.", "Total\nRs."])
            sheet.append([1,2,3,4,5,6])
            for i in range(1,totalColumns+1):
                sheet.cell(row=4, column=i).alignment = Alignment(horizontal="center", vertical="center")
                sheet.cell(row=5, column=i).alignment = Alignment(horizontal="center", vertical="center")
            sheet.print_title_rows = '4:4'
        else:
            project = projects.find_one({"_id": ObjectId(projectID)})
            projectName = project['projectName']
            village = villages.find_one({"_id": ObjectId(villageID)})
            villageName = village['villageName']
            sheet.append(["Project - " + projectName])
            rowNumber += 1
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=totalColumns)
            sheet.cell(row=3, column=1).font = Font(bold=True)
            sheet.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
            sheet.append(["Village - " + villageName])
            rowNumber += 1
            sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=totalColumns)
            sheet.cell(row=4, column=1).font = Font(bold=True)
            sheet.cell(row=4, column=1).alignment = Alignment(horizontal="center", vertical="center")
            sheet.append(["Sl. No.", "Date", "Land\nRs.", "Crop\nRs.", "Total\nRs."])
            sheet.append([1,2,3,4,5])
            for i in range(1,totalColumns+1):
                sheet.cell(row=5, column=i).alignment = Alignment(horizontal="center", vertical="center")
                sheet.cell(row=6, column=i).alignment = Alignment(horizontal="center", vertical="center")
            sheet.print_title_rows = '5:5'
        date1 = datetime.strptime(fromDate, "%d-%m-%Y").replace(day = 1)
        date2 = datetime.strptime(toDate, "%d-%m-%Y").replace(day = 1)
        months_str = calendar.month_name
        months = []
        while date1 <= date2:
            month = date1.month
            year  = date1.year
            month_str = months_str[month][0:3]
            months.append("{0} {1}".format(month_str,str(year)))
            next_month = month+1 if month != 12 else 1
            next_year = year + 1 if next_month == 1 else year
            date1 = date1.replace( month = next_month, year= next_year)
        for month in months:
            flag = 1
            if(noDataMonths == "Show"):
                sheet.append([month])
                rowNumber += 1
                sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=totalColumns)
                sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center", vertical="center")
            else:
                flag = 0
            monthLandSum = 0
            monthCropSum = 0
            monthTotalSum = 0
            for i in range(0,records_count):
                if(month == datetime.strptime(records_output[i]['date'], '%d-%m-%Y').strftime(r"%b %Y")):
                    if(flag == 0):
                        sheet.append([month])
                        rowNumber += 1
                        sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=totalColumns)
                        sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center", vertical="center")
                        flag = 1
                    rowNumber += 1
                    if(projectID == "ALL"):
                        sheet.append([i+1, records_output[i]['date'], records_output[i]['project'], records_output[i]['village'], records_output[i]['intLand'], records_output[i]['intCrop'], records_output[i]['intAmount']])
                        for col in range(5,8):
                            sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    elif(villageID == "ALL"):
                        sheet.append([i+1, records_output[i]['date'], records_output[i]['village'], records_output[i]['intLand'], records_output[i]['intCrop'], records_output[i]['intAmount']])
                        for col in range(4,7):
                            sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    else:
                        sheet.append([i+1, records_output[i]['date'], records_output[i]['intLand'], records_output[i]['intCrop'], records_output[i]['intAmount']])
                        for col in range(3,6):
                            sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
                    monthLandSum += records_output[i]['intLand']
                    monthCropSum += records_output[i]['intCrop']
                    monthTotalSum += records_output[i]['intAmount']
            if(flag == 1):
                rowNumber += 1
                if(projectID == "ALL"):
                    sheet.append([month + " Total", "", "",  "", monthLandSum, monthCropSum, monthTotalSum])
                    for col in range(5,8):
                        sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=4)
                elif(villageID == "ALL"):
                    sheet.append([month + " Total", "",  "", monthLandSum, monthCropSum, monthTotalSum])
                    for col in range(4,7):
                        sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=3)
                else:
                    sheet.append([month + " Total", "", monthLandSum, monthCropSum, monthTotalSum])
                    for col in range(3,6):
                        sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=2)
                sumRows.append(rowNumber)
                sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
        rowNumber += 1
        if(projectID == "ALL"):
            sheet.append(["Grand Total", "", "",  "", landSum, cropSum, totalSum])
            for col in range(5,8):
                sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
            sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=4)
        elif(villageID == "ALL"):
            sheet.append(["Grand Total", "",  "", landSum, cropSum, totalSum])
            for col in range(4,7):
                sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
            sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=3)
        else:
            sheet.append(["Grand Total", "", landSum, cropSum, totalSum])
            for col in range(3,6):
                sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
            sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=2)
        sumRows.append(rowNumber)
        sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
    elif(generateType == "monthlyDisbursalStatement"):
        sheet.append(["STATEMENT OF DISBURSAL (DATE WISE) - FROM " + fromDate + " TO " + toDate])
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=totalColumns)
        sheet.cell(row=2, column=1).font = Font(bold=True)
        sheet.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
        if(projectID == "ALL"):
            sheet.append(["Sl. No.", "Date", "Project",  "Village", "Name Of The Owner / Cultivator", "Land\nRs.", "Crop\nRs.", "Total\nRs.", "Cheque Number", "Serial Number In List Of Award", "Date Of Encash / Remarks"])
            sheet.append([1,2,3,4,5,6,7,8,9,10,11])
            for i in range(1,totalColumns+1):
                sheet.cell(row=3, column=i).alignment = Alignment(horizontal="center", vertical="center")
                sheet.cell(row=4, column=i).alignment = Alignment(horizontal="center", vertical="center")
            sheet.print_title_rows = '3:3'
        elif(villageID == "ALL"):
            project = projects.find_one({"_id": ObjectId(projectID)})
            projectName = project['projectName']
            sheet.append(["Project - " + projectName])
            rowNumber += 1
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=totalColumns)
            sheet.cell(row=3, column=1).font = Font(bold=True)
            sheet.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
            sheet.append(["Sl. No.", "Date", "Village", "Name Of The Owner / Cultivator", "Land\nRs.", "Crop\nRs.", "Total\nRs.", "Cheque Number", "Serial Number In List Of Award", "Date Of Encash / Remarks"])
            sheet.append([1,2,3,4,5,6,7,8,9,10])
            for i in range(1,totalColumns+1):
                sheet.cell(row=4, column=i).alignment = Alignment(horizontal="center", vertical="center")
                sheet.cell(row=5, column=i).alignment = Alignment(horizontal="center", vertical="center")
            sheet.print_title_rows = '4:4'
        else:
            project = projects.find_one({"_id": ObjectId(projectID)})
            projectName = project['projectName']
            village = villages.find_one({"_id": ObjectId(villageID)})
            villageName = village['villageName']
            sheet.append(["Project - " + projectName])
            rowNumber += 1
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=totalColumns)
            sheet.cell(row=3, column=1).font = Font(bold=True)
            sheet.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
            sheet.append(["Village - " + villageName])
            rowNumber += 1
            sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=totalColumns)
            sheet.cell(row=4, column=1).font = Font(bold=True)
            sheet.cell(row=4, column=1).alignment = Alignment(horizontal="center", vertical="center")
            sheet.append(["Sl. No.", "Date", "Name Of The Owner / Cultivator", "Land\nRs.", "Crop\nRs.", "Total\nRs.", "Cheque Number", "Serial Number In List Of Award", "Date Of Encash / Remarks"])
            sheet.append([1,2,3,4,5,6,7,8,9])
            for i in range(1,totalColumns+1):
                sheet.cell(row=5, column=i).alignment = Alignment(horizontal="center", vertical="center")
                sheet.cell(row=6, column=i).alignment = Alignment(horizontal="center", vertical="center")
            sheet.print_title_rows = '5:5'
        date1 = datetime.strptime(fromDate, "%d-%m-%Y").replace(day = 1)
        date2 = datetime.strptime(toDate, "%d-%m-%Y").replace(day = 1)
        months_str = calendar.month_name
        months = []
        while date1 <= date2:
            month = date1.month
            year  = date1.year
            month_str = months_str[month][0:3]
            months.append("{0} {1}".format(month_str,str(year)))
            next_month = month+1 if month != 12 else 1
            next_year = year + 1 if next_month == 1 else year
            date1 = date1.replace( month = next_month, year= next_year)
        for month in months:
            flag = 1
            if(noDataMonths == "Show"):
                sheet.append([month])
                rowNumber += 1
                sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=totalColumns)
                sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center", vertical="center")
            else:
                flag = 0
            monthLandSum = 0
            monthCropSum = 0
            monthTotalSum = 0
            for i in range(0,records_count):
                if(month == datetime.strptime(records_output[i]['date'], '%d-%m-%Y').strftime(r"%b %Y")):
                    if(flag == 0):
                        sheet.append([month])
                        rowNumber += 1
                        sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=totalColumns)
                        sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center", vertical="center")
                        flag = 1
                    rowNumber += 1
                    if(projectID == "ALL"):
                        sheet.append([i+1, records_output[i]['date'], records_output[i]['project'], records_output[i]['village'], records_output[i]['landOwner'], records_output[i]['intLand'], records_output[i]['intCrop'], records_output[i]['intAmount'], records_output[i]['chequeNumber'], records_output[i]['serialNumber'], records_output[i]['remarks']])
                        for col in range(6,9):
                            sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    elif(villageID == "ALL"):
                        sheet.append([i+1, records_output[i]['date'], records_output[i]['village'], records_output[i]['landOwner'], records_output[i]['intLand'], records_output[i]['intCrop'], records_output[i]['intAmount'], records_output[i]['chequeNumber'], records_output[i]['serialNumber'], records_output[i]['remarks']])
                        for col in range(5,8):
                            sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    else:
                        sheet.append([i+1, records_output[i]['date'], records_output[i]['landOwner'], records_output[i]['intLand'], records_output[i]['intCrop'], records_output[i]['intAmount'], records_output[i]['chequeNumber'], records_output[i]['serialNumber'], records_output[i]['remarks']])
                        for col in range(4,7):
                            sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
                    monthLandSum += records_output[i]['intLand']
                    monthCropSum += records_output[i]['intCrop']
                    monthTotalSum += records_output[i]['intAmount']
            if(flag == 1):
                rowNumber += 1
                if(projectID == "ALL"):
                    sheet.append([month + " Total", "", "",  "", "", monthLandSum, monthCropSum, monthTotalSum])
                    for col in range(6,9):
                        sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=5)
                elif(villageID == "ALL"):
                    sheet.append([month + " Total", "",  "", "", monthLandSum, monthCropSum, monthTotalSum])
                    for col in range(5,8):
                        sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=4)
                else:
                    sheet.append([month + " Total", "", "", monthLandSum, monthCropSum, monthTotalSum])
                    for col in range(4,7):
                        sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=3)
                sumRows.append(rowNumber)
                sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
        rowNumber += 1
        if(projectID == "ALL"):
            sheet.append(["Grand Total", "", "",  "", "", landSum, cropSum, totalSum])
            for col in range(6,9):
                sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
            sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=5)
        elif(villageID == "ALL"):
            sheet.append(["Grand Total", "", "", "", landSum, cropSum, totalSum])
            for col in range(5,8):
                sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
            sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=4)
        else:
            sheet.append(["Grand Total", "", "", landSum, cropSum, totalSum])
            for col in range(4,7):
                sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
            sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=3)
        sumRows.append(rowNumber)
        sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
    elif(generateType == "monthlyDepositAbstract" or generateType == "monthlyDisbursalAbstract"):
        if(generateType == "monthlyDepositAbstract"):
            sheet.append(["ABSTRACT OF DEPOSIT - FROM " + fromDate + " TO " + toDate])
        elif(generateType == "monthlyDisbursalAbstract"):
            sheet.append(["ABSTRACT OF DISBURSAL - FROM " + fromDate + " TO " + toDate])
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=totalColumns)
        sheet.cell(row=2, column=1).font = Font(bold=True)
        sheet.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.append(["Sl. No.", "Project", "Land\nRs.", "Crop\nRs.", "Total\nRs."])
        sheet.append([1,2,3,4,5])
        for i in range(1,totalColumns+1):
            sheet.cell(row=3, column=i).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=4, column=i).alignment = Alignment(horizontal="center", vertical="center")
        sheet.print_title_rows = '3:3'

        projects_data = projects.find()
        projects_output = [{'projectID': project['_id'], 'projectName' : project['projectName']} for project in projects_data]
        projects_count = len(projects_output)
        for i in range(0,projects_count):
            flag = 0
            for k in range(0,records_count):
                if(records_output[k]['project'] == projects_output[i]['projectName']):
                    sheet.append([i+1, records_output[k]['project'], records_output[k]['intLand'], records_output[k]['intCrop'], records_output[k]['intAmount']])
                    rowNumber += 1
                    for col in range(3,6):
                        sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
                    flag = 1
                    break
            if(flag == 0):
                sheet.append([i+1, projects_output[i]['projectName'], "-", "-", "-"])
                rowNumber += 1
                for col in range(3,6):
                    sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
        sheet.append(["Total", "",  landSum, cropSum, totalSum])
        rowNumber += 1
        for col in range(3,6):
            sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
        sumRows.append(rowNumber)
        sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=2)
        sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
    elif(generateType == "monthlyDepositStatementVillageWise" or generateType == "monthlyDisbursalStatementVillageWise"):
        if(generateType == "monthlyDepositStatementVillageWise"):
            sheet.append(["STATEMENT OF DEPOSIT (VILLAGE WISE) - FROM " + fromDate + " TO " + toDate])
        elif(generateType == "monthlyDisbursalStatementVillageWise"):
            sheet.append(["STATEMENT OF DISBURSAL (VILLAGE WISE) - FROM " + fromDate + " TO " + toDate])
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=totalColumns)
        sheet.cell(row=2, column=1).font = Font(bold=True)
        sheet.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.append(["Sl. No.", "Village", "Land\nRs.", "Crop\nRs.", "Total\nRs."])
        sheet.append([1,2,3,4,5])
        for i in range(1,totalColumns+1):
            sheet.cell(row=3, column=i).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=4, column=i).alignment = Alignment(horizontal="center", vertical="center")
        sheet.print_title_rows = '3:3'
        rowNumber = 5

        projects_data = projects.find()
        projects_output = [{'projectID': project['_id'], 'projectName' : project['projectName']} for project in projects_data]
        projects_count = len(projects_output)
        for i in range(0,projects_count):
            sheet.append(["PROJECT " + projects_output[i]['projectName']])
            sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=totalColumns)
            sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
            rowNumber += 1
            villages_data = villages.find({"projectID": projects_output[i]['projectID']})
            villages_output = [{'villageID': village['_id'], 'villageName' : village['villageName']} for village in villages_data]
            villages_count = len(villages_output)
            projectLandSum = 0
            projectCropSum = 0
            projectTotalSum = 0
            for j in range(0,villages_count):
                flag = 0
                for k in range(0,records_count):
                    if(records_output[k]['project'] == projects_output[i]['projectName'] and records_output[k]['village'] == villages_output[j]['villageName']):
                        sheet.append([j+1, records_output[k]['village'], records_output[k]['intLand'], records_output[k]['intCrop'], records_output[k]['intAmount']])
                        for col in range(3,6):
                            sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                        sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
                        rowNumber += 1
                        projectLandSum += records_output[k]['intLand']
                        projectCropSum += records_output[k]['intCrop']
                        projectTotalSum += records_output[k]['intAmount']
                        flag = 1
                        break
                if(flag == 0):
                    sheet.append([j+1, villages_output[j]['villageName'], "-", "-", "-"])
                    for col in range(3,6):
                        sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
                    rowNumber += 1
            sheet.append([projects_output[i]['projectName'] + " Total", "", projectLandSum, projectCropSum, projectTotalSum])
            sumRows.append(rowNumber)
            for col in range(3,6):
                sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
            sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=2)
            sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
            rowNumber += 1
        sheet.append(["Grand Total", "", landSum, cropSum, totalSum])
        sumRows.append(rowNumber)
        for col in range(3,6):
            sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
        sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=2)
        sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
    elif(generateType == "totalDepositStatement" or generateType == "totalDisbursalStatement"):
        if(generateType == "totalDepositStatement"):
            sheet.append(["STATEMENT OF CONSOLIDATED DEPOSIT - FROM " + fromDate + " TO " + toDate])
        elif(generateType == "totalDisbursalStatement"):
            sheet.append(["STATEMENT OF CONSOLIDATED DISBURSAL - FROM " + fromDate + " TO " + toDate])
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=totalColumns)
        sheet.cell(row=2, column=1).font = Font(bold=True)
        sheet.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
        if(generateType == "totalDepositStatement"):
            sheet.append(["Sl. No.", "Village", "Consolidated Deposit\nFrom " + previousFromDate + " To " + previousToDate, "","", "Consolidated Deposit\nFrom " + fromDate + " To "+ toDate, "","", "Consolidated Deposit\nFrom " + previousFromDate + " To " + toDate])
        elif(generateType == "totalDisbursalStatement"):
            sheet.append(["Sl. No.", "Village", "Consolidated Disbursal\nFrom " + previousFromDate + " To " + previousToDate, "","", "Consolidated Disbursal\nFrom " + fromDate + " To "+ toDate, "","", "Consolidated Disbursal\nFrom " + previousFromDate + " To " + toDate])
        sheet.row_dimensions[3].height = 36
        sheet.merge_cells(start_row=3, start_column=3, end_row=3, end_column=5)
        sheet.merge_cells(start_row=3, start_column=6, end_row=3, end_column=8)
        sheet.merge_cells(start_row=3, start_column=9, end_row=3, end_column=11)
        sheet.append(["", "", "Land\nRs.", "Crop\nRs.", "Total\nRs.", "Land\nRs.", "Crop\nRs.", "Total\nRs.", "Land\nRs.", "Crop\nRs.", "Total\nRs."])
        sheet.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
        sheet.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
        sheet.append([1,2,3,4,5,6,7,8,9,10,11])
        for i in range(1,totalColumns+1):
            sheet.cell(row=3, column=i).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=4, column=i).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=5, column=i).alignment = Alignment(horizontal="center", vertical="center")
        sheet.print_title_rows = '3:4'
        rowNumber = 6

        projects_data = projects.find()
        projects_output = [{'projectID': project['_id'], 'projectName' : project['projectName']} for project in projects_data]
        projects_count = len(projects_output)
        for i in range(0,projects_count):
            sheet.append(["PROJECT " + projects_output[i]['projectName']])
            sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=totalColumns)
            sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
            rowNumber += 1
            villages_data = villages.find({"projectID": projects_output[i]['projectID']})
            villages_output = [{'villageID': village['_id'], 'villageName' : village['villageName']} for village in villages_data]
            villages_count = len(villages_output)
            projectPastLandSum = 0
            projectPastCropSum = 0
            projectPastTotalSum = 0
            projectPresentLandSum = 0
            projectPresentCropSum = 0
            projectPresentTotalSum = 0
            projectSoFarLandSum = 0
            projectSoFarCropSum = 0
            projectSoFarTotalSum = 0
            for j in range(0,villages_count):
                flag = 0
                for k in range(0,records_count):
                    if(records_output[k]['project'] == projects_output[i]['projectName'] and records_output[k]['village'] == villages_output[j]['villageName']):
                        sheet.append([j+1, records_output[k]['village'], records_output[k]['intPastLand'], records_output[k]['intPastCrop'], records_output[k]['intPastTotal'], records_output[k]['intPresentLand'], records_output[k]['intPresentCrop'], records_output[k]['intPresentTotal'], records_output[k]['intSoFarLand'], records_output[k]['intSoFarCrop'], records_output[k]['intSoFarTotal']])
                        for col in range(3,totalColumns+1):
                            sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                        sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
                        rowNumber += 1
                        projectPastLandSum += records_output[k]['intPastLand']
                        projectPastCropSum += records_output[k]['intPastCrop']
                        projectPastTotalSum += records_output[k]['intPastTotal']
                        projectPresentLandSum += records_output[k]['intPresentLand']
                        projectPresentCropSum += records_output[k]['intPresentCrop']
                        projectPresentTotalSum += records_output[k]['intPresentTotal']
                        projectSoFarLandSum += records_output[k]['intSoFarLand']
                        projectSoFarCropSum += records_output[k]['intSoFarCrop']
                        projectSoFarTotalSum += records_output[k]['intSoFarTotal']
                        flag = 1
                        break
                if(flag == 0):
                    sheet.append([j+1, villages_output[j]['villageName'], "-", "-", "-", "-", "-", "-", "-", "-", "-"])
                    for col in range(3,totalColumns+1):
                        sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
                    rowNumber += 1
            sheet.append([projects_output[i]['projectName'] + " Total", "", projectPastLandSum, projectPastCropSum, projectPastTotalSum, projectPresentLandSum, projectPresentCropSum, projectPresentTotalSum, projectSoFarLandSum, projectSoFarCropSum, projectSoFarTotalSum])
            sumRows.append(rowNumber)
            for col in range(3,totalColumns+1):
                sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
            sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=2)
            sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
            rowNumber += 1
        sheet.append(["Grand Total", "", pastLandSum, pastCropSum, pastTotalSum, presentLandSum, presentCropSum, presentTotalSum, soFarLandSum, soFarCropSum, soFarTotalSum])
        sumRows.append(rowNumber)
        for col in range(3,totalColumns+1):
            sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
        sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=2)
        sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
    elif(generateType == "totalDepositAbstract" or generateType == "totalDisbursalAbstract"):
        if(generateType == "totalDepositAbstract"):
            sheet.append(["ABSTRACT OF CONSOLIDATED DEPOSIT - FROM " + fromDate + " TO " + toDate])
        elif(generateType == "totalDisbursalAbstract"):
            sheet.append(["ABSTRACT OF CONSOLIDATED DISBURSAL - FROM " + fromDate + " TO " + toDate])
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=totalColumns)
        sheet.cell(row=2, column=1).font = Font(bold=True)
        sheet.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
        if(generateType == "totalDepositAbstract"):
            sheet.append(["Sl. No.", "Project", "Consolidated Deposit\nFrom " + previousFromDate + " To " + previousToDate, "","", "Consolidated Deposit\nFrom " + fromDate + " To "+ toDate, "","", "Consolidated Deposit\nFrom " + previousFromDate + " To " + toDate])
        elif(generateType == "totalDisbursalAbstract"):
            sheet.append(["Sl. No.", "Project", "Consolidated Disbursal\nFrom " + previousFromDate + " To " + previousToDate, "","", "Consolidated Disbursal\nFrom " + fromDate + " To "+ toDate, "","", "Consolidated Disbursal\nFrom " + previousFromDate + " To " + toDate])
        sheet.row_dimensions[3].height = 36
        sheet.merge_cells(start_row=3, start_column=3, end_row=3, end_column=5)
        sheet.merge_cells(start_row=3, start_column=6, end_row=3, end_column=8)
        sheet.merge_cells(start_row=3, start_column=9, end_row=3, end_column=11)
        sheet.append(["", "", "Land\nRs.", "Crop\nRs.", "Total\nRs.", "Land\nRs.", "Crop\nRs.", "Total\nRs.", "Land\nRs.", "Crop\nRs.", "Total\nRs."])
        sheet.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
        sheet.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
        sheet.append([1,2,3,4,5,6,7,8,9,10,11])
        for i in range(1,totalColumns+1):
            sheet.cell(row=3, column=i).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=4, column=i).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=5, column=i).alignment = Alignment(horizontal="center", vertical="center")
        sheet.print_title_rows = '3:4'
        rowNumber = 6

        projects_data = projects.find()
        projects_output = [{'projectID': project['_id'], 'projectName' : project['projectName']} for project in projects_data]
        projects_count = len(projects_output)
        for i in range(0,projects_count):
            flag = 0
            for k in range(0,records_count):
                if(records_output[k]['project'] == projects_output[i]['projectName']):
                    sheet.append([i+1, records_output[k]['project'], records_output[k]['intPastLand'], records_output[k]['intPastCrop'], records_output[k]['intPastTotal'], records_output[k]['intPresentLand'], records_output[k]['intPresentCrop'], records_output[k]['intPresentTotal'], records_output[k]['intSoFarLand'], records_output[k]['intSoFarCrop'], records_output[k]['intSoFarTotal']])
                    for col in range(3,totalColumns+1):
                        sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
                    rowNumber += 1
                    flag = 1
                    break
            if(flag == 0):
                sheet.append([i+1, projects_output[i]['projectName'], "-", "-", "-", "-", "-", "-", "-", "-", "-"])
                for col in range(3,totalColumns+1):
                    sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
                rowNumber += 1
        sheet.append(["Grand Total", "", pastLandSum, pastCropSum, pastTotalSum, presentLandSum, presentCropSum, presentTotalSum, soFarLandSum, soFarCropSum, soFarTotalSum])
        for col in range(3,totalColumns+1):
            sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
        sumRows.append(rowNumber)
        sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=2)
        sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
    elif(generateType == "balanceSheetStatement"):
        sheet.append(["STATEMENT OF BALANCE SHEET FOR DEPOSIT & DISBURSAL - FROM " + fromDate + " TO " + toDate])
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=totalColumns)
        sheet.cell(row=2, column=1).font = Font(bold=True)
        sheet.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.append(["Sl. No.", "Village", "Grand Total Deposit\nFrom " + fromDate + " To " + toDate, "","", "Grand Total Disbursal\nFrom " + fromDate + " To "+ toDate, "","", "Balance"])
        sheet.row_dimensions[3].height = 36
        sheet.merge_cells(start_row=3, start_column=3, end_row=3, end_column=5)
        sheet.merge_cells(start_row=3, start_column=6, end_row=3, end_column=8)
        sheet.merge_cells(start_row=3, start_column=9, end_row=3, end_column=11)
        sheet.append(["", "", "Land\nRs.", "Crop\nRs.", "Total\nRs.", "Land\nRs.", "Crop\nRs.", "Total\nRs.", "Land\nRs.", "Crop\nRs.", "Total\nRs."])
        sheet.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
        sheet.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
        sheet.append([1,2,3,4,5,6,7,8,9,10,11])
        for i in range(1,totalColumns+1):
            sheet.cell(row=3, column=i).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=4, column=i).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=5, column=i).alignment = Alignment(horizontal="center", vertical="center")
        sheet.print_title_rows = '3:4'
        rowNumber = 6

        projects_data = projects.find()
        projects_output = [{'projectID': project['_id'], 'projectName' : project['projectName']} for project in projects_data]
        projects_count = len(projects_output)
        for i in range(0,projects_count):
            sheet.append(["PROJECT " + projects_output[i]['projectName']])
            sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=totalColumns)
            sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
            rowNumber += 1
            villages_data = villages.find({"projectID": projects_output[i]['projectID']})
            villages_output = [{'villageID': village['_id'], 'villageName' : village['villageName']} for village in villages_data]
            villages_count = len(villages_output)
            projectDepositLandSum = 0
            projectDepositCropSum = 0
            projectDepositTotalSum = 0
            projectDisbursalLandSum = 0
            projectDisbursalCropSum = 0
            projectDisbursalTotalSum = 0
            projectBalanceLandSum = 0
            projectBalanceCropSum = 0
            projectBalanceTotalSum = 0

            for j in range(0,villages_count):
                flag = 0
                for k in range(0,records_count):
                    if(records_output[k]['project'] == projects_output[i]['projectName'] and records_output[k]['village'] == villages_output[j]['villageName']):
                        sheet.append([j+1, records_output[k]['village'], records_output[k]['intDepositLand'], records_output[k]['intDepositCrop'], records_output[k]['intDepositTotal'], records_output[k]['intDisbursalLand'], records_output[k]['intDisbursalCrop'], records_output[k]['intDisbursalTotal'], records_output[k]['intBalanceLand'], records_output[k]['intBalanceCrop'], records_output[k]['intBalanceTotal']])
                        for col in range(3,totalColumns+1):
                            sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                        sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
                        rowNumber += 1
                        projectDepositLandSum += records_output[k]['intDepositLand']
                        projectDepositCropSum += records_output[k]['intDepositCrop']
                        projectDepositTotalSum += records_output[k]['intDepositTotal']
                        projectDisbursalLandSum += records_output[k]['intDisbursalLand']
                        projectDisbursalCropSum += records_output[k]['intDisbursalCrop']
                        projectDisbursalTotalSum += records_output[k]['intDisbursalTotal']
                        projectBalanceLandSum += records_output[k]['intBalanceLand']
                        projectBalanceCropSum += records_output[k]['intBalanceCrop']
                        projectBalanceTotalSum += records_output[k]['intBalanceTotal']
                        flag = 1
                        break
                if(flag == 0):
                    sheet.append([j+1, villages_output[j]['villageName'], "-", "-", "-", "-", "-", "-", "-", "-", "-"])
                    for col in range(3,totalColumns+1):
                        sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
                    rowNumber += 1
            sheet.append([projects_output[i]['projectName'] + " Total", "", projectDepositLandSum, projectDepositCropSum, projectDepositTotalSum, projectDisbursalLandSum, projectDisbursalCropSum, projectDisbursalTotalSum, projectBalanceLandSum, projectBalanceCropSum, projectBalanceTotalSum])
            sumRows.append(rowNumber)
            for col in range(3,totalColumns+1):
                sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
            sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=2)
            sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
            rowNumber += 1
        sheet.append(["Grand Total", "", depositLandSum, depositCropSum, depositTotalSum, disbursalLandSum, disbursalCropSum, disbursalTotalSum, balanceLandSum, balanceCropSum, balanceTotalSum])
        sumRows.append(rowNumber)
        for col in range(3,totalColumns+1):
            sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
        sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=2)
        sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
    elif(generateType == "balanceSheetAbstract"):
        sheet.append(["ABSTRACT OF BALANCE SHEET FOR DEPOSIT & DISBURSAL - FROM " + fromDate + " TO " + toDate])
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=totalColumns)
        sheet.cell(row=2, column=1).font = Font(bold=True)
        sheet.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.append(["Sl. No.", "Project", "Grand Total Deposit\nFrom " + fromDate + " To " + toDate, "","", "Grand Total Disbursal\nFrom " + fromDate + " To "+ toDate, "","", "Balance"])
        sheet.row_dimensions[3].height = 36
        sheet.merge_cells(start_row=3, start_column=3, end_row=3, end_column=5)
        sheet.merge_cells(start_row=3, start_column=6, end_row=3, end_column=8)
        sheet.merge_cells(start_row=3, start_column=9, end_row=3, end_column=11)
        sheet.append(["", "", "Land\nRs.", "Crop\nRs.", "Total\nRs.", "Land\nRs.", "Crop\nRs.", "Total\nRs.", "Land\nRs.", "Crop\nRs.", "Total\nRs."])
        sheet.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
        sheet.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
        sheet.append([1,2,3,4,5,6,7,8,9,10,11])
        for i in range(1,totalColumns+1):
            sheet.cell(row=3, column=i).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=4, column=i).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=5, column=i).alignment = Alignment(horizontal="center", vertical="center")
        sheet.print_title_rows = '3:4'
        rowNumber = 6

        projects_data = projects.find()
        projects_output = [{'projectID': project['_id'], 'projectName' : project['projectName']} for project in projects_data]
        projects_count = len(projects_output)
        for i in range(0,projects_count):
            flag = 0
            for k in range(0,records_count):
                if(records_output[k]['project'] == projects_output[i]['projectName']):
                    sheet.append([i+1, records_output[k]['project'], records_output[k]['intDepositLand'], records_output[k]['intDepositCrop'], records_output[k]['intDepositTotal'], records_output[k]['intDisbursalLand'], records_output[k]['intDisbursalCrop'], records_output[k]['intDisbursalTotal'], records_output[k]['intBalanceLand'], records_output[k]['intBalanceCrop'], records_output[k]['intBalanceTotal']])
                    for col in range(3,totalColumns+1):
                        sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                    sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
                    rowNumber += 1
                    flag = 1
                    break
            if(flag == 0):
                sheet.append([i+1, projects_output[i]['projectName'], "-", "-", "-", "-", "-", "-", "-", "-", "-"])
                for col in range(3,totalColumns+1):
                    sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
                sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
                rowNumber += 1
        sheet.append(["Grand Total", "", depositLandSum, depositCropSum, depositTotalSum, disbursalLandSum, disbursalCropSum, disbursalTotalSum, balanceLandSum, balanceCropSum, balanceTotalSum])
        for col in range(3,totalColumns+1):
            sheet.cell(row=rowNumber, column=col).number_format = currencyFormat
        sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=2)
        sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")
        sumRows.append(rowNumber)
        sheet.merge_cells(start_row=rowNumber, start_column=1, end_row=rowNumber, end_column=2)
        sheet.cell(row=rowNumber, column=1).alignment = Alignment(horizontal="center")

    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    for row in sheet.iter_rows(min_row=1, max_col=totalColumns, max_row=rowNumber):
        for cell in row:
            cell.border = border
            cell.alignment =  cell.alignment.copy(wrapText=True)
            cell.font = Font(size=14)
            if(cell.value == 0):
                cell.value = "-"
                cell.alignment = Alignment(horizontal="right")
    sheet.cell(row=2, column=1).font = Font(bold=True, size=16)
    sumRows.append(rowNumber)
    for sumRow in sumRows:
        for no in range(1,totalColumns+1):
            sheet.cell(row=sumRow, column=no).font = Font(bold=True, size=14)
    for idx, col in enumerate(sheet.columns, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].auto_size = True

    basedir = os.path.abspath(os.path.dirname(__file__))
    fileDir = os.path.join(basedir, 'static/')
    fileName = "STATEMENT OF SRPL PROJECTS IOCL THIRUVALLUR.xlsx"
    wb.save(fileDir + fileName)
    return send_from_directory(fileDir, fileName, as_attachment=True)

@app.route("/config")
def config():
    session_id = request.cookies.get('session_id')
    if(session_id != "mjeIJPsatvvs"):
        return redirect(url_for('login'))
        
    projects_data = projects.find()
    projects_output = [{'projectID': project['_id'], 'projectName' : project['projectName']} for project in projects_data]
    projects_count = len(projects_output)
    return render_template('config.html', projects_output=projects_output, projects_count=projects_count, display_villages=False)

@app.route("/getVillages", methods=['POST'])
def getVillages():
    session_id = request.cookies.get('session_id')
    if(session_id != "mjeIJPsatvvs"):
        return redirect(url_for('login'))
        
    projects_data = projects.find()
    projects_output = [{'projectID': project['_id'], 'projectName' : project['projectName']} for project in projects_data]
    projects_count = len(projects_output)

    villages_data = villages.find({"projectID": ObjectId(request.form['projectID'])})
    villages_output = [{'villageID': village['_id'], 'villageName' : village['villageName']} for village in villages_data]
    villages_count = len(villages_output)

    projectName = ""
    for project in projects_output:
        if(ObjectId(request.form['projectID']) == project['projectID']):
            projectName = project['projectName']
            break

    return render_template('config.html', projects_output=projects_output, projects_count=projects_count, villages_output=villages_output, villages_count=villages_count, display_villages=True, projectID=request.form['projectID'], projectIDObj=ObjectId(request.form['projectID']), projectName=projectName)

@app.route("/getVillagesOfProject/<projectID>")
def getVillagesOfProject(projectID):
    session_id = request.cookies.get('session_id')
    if(session_id != "mjeIJPsatvvs"):
        return redirect(url_for('login'))
        
    villages_data = villages.find({"projectID": ObjectId(projectID)})
    villages_output = [{'villageID': str(village['_id']), 'villageName' : village['villageName']} for village in villages_data]
    villages_count = len(villages_output)
    response = {
        "villages_output": villages_output,
        "villages_count": villages_count
    }
    return jsonify(response)

@app.route("/login")
def login():
    return render_template('login.html', err=False)

@app.route("/authenticate", methods=['POST'])
def authenticate():
    username = request.form['username']
    password = request.form['password']
    if((username == "rprapagar-Rev.Acct.") and (password == "Sriramajayam@123")):
        response = make_response(redirect(url_for('statements')))
        expire_date = datetime.now()
        expire_date = expire_date + timedelta(days=1)
        response.set_cookie('session_id', 'mjeIJPsatvvs', expires=expire_date)
        return response
    else:
        return render_template('login.html', err=True)

@app.route("/logout")
def logout():
    response = make_response(redirect(url_for('login')))
    response.set_cookie('session_id', '')
    return response

@app.route("/generateExcelPage")
def generateExcelPage():
    projects_data = projects.find()
    projects_output = [{'projectID': str(project['_id']), 'projectName' : project['projectName']} for project in projects_data]
    projects_count = len(projects_output)
    return render_template('generateExcel.html', err=False, projects_output=projects_output, projects_count=projects_count)

@app.route("/isChequeNumberExists/<chequeNumber>")
def isChequeNumberExists(chequeNumber):
    session_id = request.cookies.get('session_id')
    if(session_id != "mjeIJPsatvvs"):
        return redirect(url_for('login'))
        
    records_data = records.find({"chequeNumber": chequeNumber})
    records_output = [{'recordID': str(record['_id'])} for record in records_data]
    records_count = len(records_output)
    response = {}
    if(records_count > 0):
        response = {
            "isChequeNumberExists": "Y"
        }
    else:
        response = {
            "isChequeNumberExists": "N"
        }
    return jsonify(response)

if __name__ == '__main__':
    app.run(debug=True)